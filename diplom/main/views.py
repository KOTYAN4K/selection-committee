import os
import string

import openpyxl
from django.http import HttpResponse, FileResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import TemplateView, ListView, FormView, CreateView
from openpyxl.workbook import Workbook

from main.admin import StudentResource
from main.forms import ApplicantShortForm, FieldSelectionForm
from main.models import Department, Applicant, Parent, School, ApplicantAdmissionView, Document, Admission
from main.utils import parse_schools, fill_template


# Create your views here.
class IndexView(TemplateView):
    template_name = 'main/index.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['departments'] = Department.objects.all()
        return context


class TestFormView(CreateView):
    template_name = 'main/form.html'
    form_class = ApplicantShortForm
    success_url = reverse_lazy('home')


def download_table(request):
    # Получите данные из базы данных
    queryset = Applicant.objects.all()

    # Создайте новый файл Excel
    wb = Workbook()
    ws = wb.active
    ws.title = 'Студенты'

    ws.append([
        '№ п/п',
        'ФИО',
        'пол',
        'Кол "5"',
        'Кол "4"',
        'Кол "3"',
        'Снилс',
        'инн',
        'средн балл',
        'оригинал/копия',
        'внебюджет',
        'Документы | забрали',
        'Получил ли расписку',
        'школа',
        'год окончания',
        'ФИС',
        'номер заявления',
        'Электронное заявление',
        'Дата ЭЗ',
        'Дата и время проведения испытания',
        'Статус',
        'Вступительный экзамен',
        'Дата рождения',
        'Личный телефон',
        'Мама',
        'Телефон мамы',
        'Папа',
        'Телефон папы',
        'номер паспорта',
        'Кем выдан',
        'Дата выдачи'
    ])

    # Добавьте данные из базы данных в файл Excel
    for obj in queryset:
        ws.append([obj.pk,
                   f'{obj.first_name} {obj.last_name} {obj.patronymic}',  # ФИО
                   f'{obj.gender}',  # Пол
                   f'{obj.student.number_of_5}',
                   f'{obj.student.number_of_4}',
                   f'{obj.student.number_of_3}',
                   f'{obj.document.SNILS}',
                   f'{obj.document.INN}',
                   f'{obj.student.average_score}',
                   f'{obj.student.original_or_copy}',
                   f'{obj.student.out_of_budget}',
                   f'{obj.student.documents_collected}',
                   f'{obj.student.received_receipt}',
                   f'{obj.school}',
                   f'{obj.graduation_date}',
                   f'{obj.document.FIS}',
                   f'{str(obj.id).zfill(5)}',
                   f'{obj.student.application_in_gov_services}',
                   f'',
                   f'',
                   f'{obj.student.application_status}',
                   f'{obj.student.internal_exam}',
                   f'{obj.birth_date}',
                   f'{obj.phone}',
                   f'{obj.parents.mother_full_name}',
                   f'{obj.parents.mother_phone}',
                   f'{obj.parents.father_full_name}',
                   f'{obj.parents.father_phone}',
                   f'{obj.document.passport_number}',
                   f'{obj.document.issued_by}',
                   f'{obj.document.issue_date}'
                   ])

    # Создаём HTTP-ответ для скачивания файла
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Applicants.xlsx'

    # Сохраняем файл Excel в HTTP-ответ
    wb.save(response)
    return response


def download_document(request):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # Define text file name
    filename = 'tipovaya-forma_soglasie.doc'
    # Define the full file path
    filepath = BASE_DIR + '\\static\\documents\\' + filename
    # Open the file for reading content
    response = FileResponse(open(filepath, 'rb'))
    response['Content-Disposition'] = 'attachment; filename="your_file.doc"'
    # Return the response value
    return response


def update_schools(request):
    parse_schools()
    return redirect('home')


def generate_document(request, person_id):
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    person = get_object_or_404(Applicant, id=person_id)
    template_path = BASE_DIR + '\\static\\documents\\' + 'ZayavlenieAbiturienta.docx'
    filled_doc = fill_template(person.id, template_path)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
    response['Content-Disposition'] = f'attachment; filename="document.docx"'
    filled_doc.save(response)

    return response


def normalize(text):
    return ''.join(char for char in text if char not in string.punctuation).lower()


def autocomplete(request):
    if 'term' in request.GET:
        term_request = normalize(request.GET.get('term'))

        qs = School.objects.filter(name__iregex=term_request)
        title = []
        for school in qs:
            title.append(school.name)
        return JsonResponse(title, safe=False)

    render(request, 'main/form.html')


def export_data(request):
    if request.method == 'GET':
        form = ApplicantShortForm()

        return render(request, 'admin/export_data.html')

    if request.method == 'POST':
        queryset = Applicant.objects.all()

        # Создайте новый файл Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Студенты'

        ws.append([
            '№ п/п',
            'ФИО',
            'пол',
            'Кол "5"',
            'Кол "4"',
            'Кол "3"',
            'Снилс',
            'инн',
            'средн балл',
            'оригинал/копия',
            'внебюджет',
            'Документы | забрали',
            'Получил ли расписку',
            'школа',
            'год окончания',
            'ФИС',
            'номер заявления',
            'Электронное заявление',
            'Дата ЭЗ',
            'Дата и время проведения испытания',
            'Статус',
            'Вступительный экзамен',
            'Дата рождения',
            'Личный телефон',
            'Мама',
            'Телефон мамы',
            'Папа',
            'Телефон папы',
            'номер паспорта',
            'Кем выдан',
            'Дата выдачи'
        ])

        # Добавьте данные из базы данных в файл Excel
        for obj in queryset:
            parents = Parent.objects.get(student_id=obj.pk)
            ws.append([obj.pk,
                       f'{obj.first_name} {obj.last_name} {obj.patronymic}',  # ФИО
                       f'{obj.gender}',  # Пол
                       'Кол "5"',
                       'Кол "4"',
                       'Кол "3"',
                       'Снилс',
                       'инн',
                       'средн балл',
                       'оригинал/копия',
                       'внебюджет',
                       'Документы | забрали',
                       'Получил ли расписку',
                       f'{obj.school}',
                       f'{obj.graduation_date}',
                       'ФИС',
                       'номер заявления',
                       'Электронное заявление',
                       'Дата ЭЗ',
                       'Дата и время проведения испытания',
                       'Статус',
                       'Вступительный экзамен',
                       'Дата рождения',
                       'Личный телефон',
                       f'{parents.mother_full_name}',
                       f'{parents.mother_phone}',
                       f'{parents.father_full_name}',
                       f'{parents.father_phone}',
                       'номер паспорта',
                       'Кем выдан',
                       'Дата выдачи'
                       ])

        # Создаём HTTP-ответ для скачивания файла
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Applicants.xlsx'

        # Сохраняем файл Excel в HTTP-ответ
        wb.save(response)
        return response


def export_students(request):
    dataset = StudentResource().export()
    response = HttpResponse(dataset.xlsx,
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="students.xlsx"'
    return response


def export_to_excel(request):
    if request.method == 'POST':
        form = FieldSelectionForm(request.POST)
        if form.is_valid():
            selected_fields = {
                'applicant': form.cleaned_data.get('APPLICANT_CHOICES', []),
                'document': form.cleaned_data.get('DOCUMENT_CHOICES', []),
                'parent': form.cleaned_data.get('PARENT_CHOICES', []),
                'admission': form.cleaned_data.get('ADMISSION_CHOICES', [])
            }

            # Create a workbook and add a worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Applicant Data"

            # Add headers
            row_num = 1
            col_num = 1
            headers = [('ID', 'ID')]  # Add 'ID' as the first column
            verbose_names = ['ID']
            order = ['last_name', 'first_name', 'patronymic', 'gender', 'birth_date', 'email', 'phone', 'address',
                     'school', 'graduation_date', 'education', 'consent', 'SNILS',
                     'INN', 'passport_number', 'certificate', 'FIS', 'mother_full_name',
                     'mother_phone', 'father_full_name', 'father_phone', 'admission_date', 'number_of_5', 'number_of_4',
                     'number_of_3', 'average_score', 'internal_exam', 'application_status', 'original_or_copy',
                     'out_of_budget', 'received_receipt', 'internal_exam_conducted', 'documents_collected',
                     'application_in_gov_services']

            def sort_fields(fields):
                return sorted(fields, key=lambda x: order.index(x[0]) if x[0] in order else len(order))

            # Map selected fields to their verbose names
            models_map = {
                'applicant': Applicant,
                'document': Document,
                'parent': Parent,
                'admission': Admission,
            }

            for model, fields in selected_fields.items():
                sorted_fields = sort_fields([(field, field) for field in fields])
                for field, _ in sorted_fields:
                    headers.append((model, field))
                    verbose_name = models_map[model]._meta.get_field(field).verbose_name
                    verbose_names.append(verbose_name)
                    ws.cell(row=row_num, column=col_num + 1, value=verbose_name)  # Shift by 1 for ID
                    col_num += 1

            # Add headers for ID
            ws.cell(row=row_num, column=1, value='ID')

            # Add data
            for idx, obj in enumerate(ApplicantAdmissionView.objects.all(), start=1):
                row_num += 1
                col_num = 1
                # Add ID
                ws.cell(row=row_num, column=col_num, value=idx)
                col_num += 1
                for model, field in headers[1:]:  # Skip the first ID header
                    related_obj = getattr(obj, model)
                    value = getattr(related_obj, field)
                    if callable(value):
                        value = value()
                    if isinstance(value, bool):
                        value = 'Да' if value else 'Нет'
                    ws.cell(row=row_num, column=col_num, value=value)
                    col_num += 1

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

            # Adjust row heights
            for row in ws.iter_rows():
                max_height = 15  # Default height
                for cell in row:
                    if cell.value:
                        cell_height = (len(str(cell.value)) // 20) * 5
                        if cell_height > max_height:
                            max_height = cell_height
                ws.row_dimensions[cell.row].height = max_height

            # Save the workbook
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename=applicant_data.xlsx'
            wb.save(response)
            return response
    else:
        form = FieldSelectionForm()

    return render(request, 'admin/export_data.html', {'form': form})


def page_not_found(request, exception):
    return HttpResponse("404 NOT FOUND")
